import json
from time import sleep

import openpyxl
import pandas as pd
import requests
from openpyxl.styles import Alignment, Font


def read_cve_with_pandas(file_path):
  try:
    # header=None 表示文件没有表头，第一行就是数据
    df = pd.read_excel(file_path, header=None)

    # 提取第一列 (索引为0)，去掉空值(NaN)，转为字符串并去除两端空格，最后转为 list
    cve_list = df[0].dropna().astype(str).str.strip().tolist()

    # 过滤掉不以 CVE- 开头的脏数据
    cve_list = [cve for cve in cve_list if cve.upper().startswith("CVE-")]

    return cve_list
  except Exception as e:
    print(f"[-] 读取失败: {e}")
    return


def write_cve_to_excel(data_list, output_file="cve_results.xlsx"):
  """
  将收集到的 CVE 漏洞信息写入 Excel (xlsx) 文件中。
  xlsx 格式原生支持并默认使用 UTF-8 编码存储文本。

  :param data_list: 包含漏洞信息的字典列表
  :param output_file: 保存的文件名或路径
  """
  try:
    # 1. 创建一个新的 Excel 工作簿和工作表
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "漏洞信息收集"

    # 2. 定义并写入表头
    headers = ["漏洞编号", "漏洞描述", "CVSS 评分", "受影响的组件"]
    sheet.append(headers)

    # 加粗表头并居中
    for col_num, header in enumerate(headers, 1):
      cell = sheet.cell(row=1, column=col_num)
      cell.font = Font(bold=True)
      cell.alignment = Alignment(horizontal="center", vertical="center")

    # 3. 遍历数据并逐行写入
    for item in data_list:
      cve_id = item.get("cve_id", "未知")
      description = item.get("description", "暂无描述")
      cvss_score = item.get("cvss_score", "未知")
      components = item.get("affected_components", [])

      # 如果受影响的组件是一个列表，用换行符拼接成一个字符串
      if isinstance(components, list):
        components_str = "\n".join(components)
      else:
        components_str = str(components)

      # 写入当前行
      row_data = [cve_id, description, cvss_score, components_str]
      sheet.append(row_data)

    # 4. 美化表格：调整列宽和设置自动换行
    sheet.column_dimensions['A'].width = 18  # 漏洞编号
    sheet.column_dimensions['B'].width = 60  # 漏洞描述 (较宽)
    sheet.column_dimensions['C'].width = 15  # CVSS 评分
    sheet.column_dimensions['D'].width = 60  # 受影响的组件 (较宽)

    # 从第二行开始遍历所有数据行，设置文本自动换行和垂直居中
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=4):
      for cell in row:
        cell.alignment = Alignment(wrapText=True, vertical="center")

    # 5. 保存文件
    workbook.save(output_file)
    print(f"\n[+] 成功将 {len(data_list)} 条漏洞信息保存至 '{output_file}' (UTF-8 编码)")

  except Exception as e:
    print(f"\n[-] 写入 Excel 文件时发生错误: {e}")


def get_cve_info(cve_id):
  """
  通过 NVD API 获取指定 CVE 的信息，并返回格式化后的字典数据。
  :param cve_id: CVE编号，例如 "CVE-2015-2483"
  :return: 包含漏洞信息的字典 (dict)，如果失败则返回 None
  """
  url = f"https://services.nvd.nist.gov/rest/json/cves/2.0?cveId={cve_id}"

  # NVD API 建议加上 User-Agent
  headers = {
      "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko)",
      "apiKey": "e2ddf669-86be-41ba-9f18-7dbf12c6e793"
  }

  try:
    print(f"[*] 正在请求 NVD 接口获取 {cve_id} 的数据...")
    response = requests.get(url, headers=headers, timeout=15)

    if response.status_code != 200:
      print(f"[-] {cve_id} 请求失败！状态码: {response.status_code}")
      return None

    data = response.json()

    # 检查是否返回了漏洞数据
    if not data.get("vulnerabilities") or len(data["vulnerabilities"]) == 0:
      print(f"[-] 未找到关于 {cve_id} 的信息，可能是该 CVE 编号不存在或尚未公开。")
      return None

    # 提取第一个结果
    cve_node = data["vulnerabilities"][0]["cve"]

    # 1. 提取描述信息 (优先获取英文)
    description_text = "暂无描述"
    for desc in cve_node.get("descriptions", []):
      if desc.get("lang") == "en":
        description_text = desc.get("value")
        break

    # 2. 提取 CVSS 评分 (优先取 V3.1 -> V3.0 -> V2)
    metrics = cve_node.get("metrics", {})
    cvss_score = "未知"
    cvss_severity = "未知"

    if "cvssMetricV31" in metrics:
      metric_data = metrics["cvssMetricV31"][0]
      cvss_score = metric_data["cvssData"].get("baseScore", "未知")
      cvss_severity = metric_data["cvssData"].get("baseSeverity", "未知")
    elif "cvssMetricV30" in metrics:
      metric_data = metrics["cvssMetricV30"][0]
      cvss_score = metric_data["cvssData"].get("baseScore", "未知")
      cvss_severity = metric_data["cvssData"].get("baseSeverity", "未知")
    elif "cvssMetricV2" in metrics:
      metric_data = metrics["cvssMetricV2"][0]
      cvss_score = metric_data["cvssData"].get("baseScore", "未知")
      cvss_severity = metric_data.get("baseSeverity", "未知")

    # 将评分和严重等级合并为一个字符串，例如 "5.0 (MEDIUM)"
    if cvss_score != "未知" or cvss_severity != "未知":
      formatted_cvss = f"{cvss_score} ({cvss_severity})"
    else:
      formatted_cvss = "未知"

    # 3. 提取受影响的组件 (CPE)
    affected_components = []
    configurations = cve_node.get("configurations", [])
    for config in configurations:
      for node in config.get("nodes", []):
        for cpe_match in node.get("cpeMatch", []):
          if cpe_match.get("vulnerable"):  # 只提取被标记为脆弱的组件
            affected_components.append(cpe_match.get("criteria"))

    # 去重
    affected_components = list(set(affected_components))

    print(f"[+] {cve_id} 数据解析成功！")

    # 4. 组装成目标字典格式并返回
    result_data = {
        "cve_id": cve_node.get("id", cve_id),
        "description": description_text,
        "cvss_score": formatted_cvss,
        "affected_components": affected_components
    }

    return result_data

  except requests.exceptions.RequestException as e:
    print(f"[-] {cve_id} 网络请求异常: {e}")
    return None
  except KeyError as e:
    print(f"[-] {cve_id} 数据解析异常，找不到预期字段: {e}")
    return None


if __name__ == "__main__":
  # vuln_ids = read_cve_with_pandas("./data/vuln_id_2015-2025.xlsx")
  # print("Read Ok!")
  vuln_ids = ["CVE-2022-27965", "CVE-2022-33035"]

  result = []
  skipped_vuln_id = []

  i = 0
  for vuln_id in vuln_ids:
    if vuln_id.startswith("CVE-"):
      info = get_cve_info(vuln_id)

      if info is None:
        skipped_vuln_id.append(vuln_id)
        info = {
            "cve_id": vuln_id,
            "description": "",
            "cvss_score": "",
            "affected_components": ""
        }

      result.append(info)
    i += 1
    print(i, vuln_id)
    sleep(1)
  print("Query Ok!")
  print(result)

  # write_cve_to_excel(
  #     result, output_file="result/vuln_info_fetch/cve_results.xlsx")
  # print("Write Ok!")

  # with open("result/vuln_info_fetch/skipped_vuln_id.json", "w", encoding="utf-8") as f:
  #   f.write(json.dumps(skipped_vuln_id))
