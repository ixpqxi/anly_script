import os
import time

import dotenv
import openpyxl
from openai import OpenAI

dotenv.load_dotenv()


def translate_cve_descriptions(file_path, api_key, base_url="https://api.openai.com/v1", model_name="gpt-3.5-turbo"):
  """
  读取 Excel 文件，调用大模型将第二列的英文漏洞描述翻译为中文，并覆盖原单元格。

  :param file_path: xlsx 文件的路径
  :param api_key: 大模型 API Key
  :param base_url: 大模型 API 请求基础地址 (兼容 OpenAI 格式)
  :param model_name: 调用的模型名称
  """
  # 1. 初始化 LLM 客户端
  client = OpenAI(
      api_key=api_key,
      base_url=base_url
  )

  # 精心设计的提示词 (Prompt)，确保专业且只输出结果
  system_prompt = """你是一个网络安全领域的专业翻译助手。
请将用户输入的 CVE 漏洞英文描述翻译成地道、专业的中文。
请注意保持网络安全专业术语的准确性（例如：remote attackers翻译为远程攻击者，arbitrary code翻译为任意代码，memory corruption翻译为内存破坏等）。
**严格要求：请直接输出翻译后的纯中文结果，不要包含任何前缀、后缀、引号或多余的解释。**"""

  try:
    print(f"[*] 正在加载 Excel 文件: {file_path}")
    workbook = openpyxl.load_workbook(file_path)
    # 获取第一张表
    sheet = workbook.worksheets[0]

    # 获取最大行数
    max_row = sheet.max_row
    print(f"[*] 表格共有 {max_row} 行数据（包含表头）。开始翻译...")

    # 2. 从第2行开始遍历（跳过第1行的表头）
    for row in range(2, max_row + 1):
      # 获取第二列（即 B 列，漏洞描述）的单元格
      cell = sheet.cell(row=row, column=2)
      original_text = cell.value

      # 检查单元格是否有内容，并且是否是英文字符串
      if original_text and isinstance(original_text, str):
        print(f"\n[+] 正在翻译第 {row} 行...")
        print(f"  原文本前缀: {original_text[:50]}...")

        try:
          # 调用大模型进行翻译
          response = client.chat.completions.create(
              model=model_name,
              messages=[
                  {"role": "system", "content": system_prompt},
                  {"role": "user", "content": original_text}
              ],
              temperature=0.3,  # 较低的温度使得翻译结果更稳定、准确
          )

          # 提取翻译结果并去除首尾空白字符
          translated_text = response.choices[0].message.content.strip()

          # 3. 将翻译结果写回原单元格，覆盖原英文
          cell.value = translated_text
          print(f"  翻译完成: {translated_text[:30]}...")

          # 稍作延时，防止触发 API 速率限制 (Rate Limit)
          time.sleep(1)

        except Exception as api_err:
          print(f"  [-] 第 {row} 行翻译失败 (API 异常): {api_err}")
          # 翻译失败时跳过，保留原英文
          continue

    # 4. 保存覆盖写入后的 Excel 文件
    # 注意：保存时请确保 Excel 文件在你的电脑上处于关闭状态，否则会报错 PermissionError
    workbook.save(file_path)
    print(f"\n[+] 所有翻译已完成，结果已保存覆盖至 '{file_path}'")

  except Exception as e:
    print(f"[-] 处理 Excel 文件时发生错误: {e}")


# ==========================================
# 模拟调用代码
# ==========================================
if __name__ == "__main__":
  # 你的 Excel 文件路径
  target_excel = "result/vuln_info_fetch/cve_results.xlsx"

  # === 大模型配置区域 ===
  # 以下使用 DeepSeek 的 API 作为示例 (性价比高)，你也可以换成其他的。
  # 如果使用 OpenAI官方，只需填 API_KEY 即可，Base_url 留空使用默认值。

  API_KEY = os.getenv("DEEPSEEK_KEY")
  BASE_URL = os.getenv("DEEPSEEK_URL")
  # 替换为对应的模型名，如 "gpt-3.5-turbo", "qwen-turbo"等
  MODEL = os.getenv("MODEL")

  translate_cve_descriptions(
      file_path=target_excel,
      api_key=API_KEY,
      base_url=BASE_URL,
      model_name=MODEL
  )
